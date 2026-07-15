import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

path = "/home/claude/day3/output/Customer_RFM_Segments.xlsx"
wb = openpyxl.load_workbook(path)

if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws = wb.create_sheet("Dashboard", 0)
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2

title_fill = PatternFill("solid", fgColor="2E5266")
ws["B2"] = "Day 3 — Customer Segmentation (RFM Analysis) Dashboard"
ws["B2"].font = Font(size=16, bold=True, color="FFFFFF")
ws.merge_cells("B2:N2")
ws["B2"].fill = title_fill
ws["B2"].alignment = Alignment(vertical="center", indent=1)
ws.row_dimensions[2].height = 30

ws["B3"] = "Project: Algorithmic Credit Scoring | Source: Master_Finance_Data_Cleaned.xlsx (Day 1) | Customers segmented: 801"
ws["B3"].font = Font(size=10, italic=True, color="666666")
ws.merge_cells("B3:N3")

images = [
    ("01_segment_sizes.png", "B5"),
    ("05_rf_heatmap.png", "H5"),
    ("02_rfm_scatter.png", "B24"),
    ("03_default_rate_by_segment.png", "H24"),
    ("04_creditscore_by_segment.png", "H43"),
]
for fname, cell in images:
    img = XLImage(f"/home/claude/day3/charts/{fname}")
    img.width = 430
    img.height = 260
    ws.add_image(img, cell)

wb.save(path)
print("Dashboard added.", wb.sheetnames)
