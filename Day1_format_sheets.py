import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

path = "/home/claude/day1/output/Master_Finance_Data_Cleaned.xlsx"
wb = openpyxl.load_workbook(path)

header_fill = PatternFill("solid", fgColor="2E5266")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
body_font = Font(name="Arial", size=10)

data_sheets = ["Cleaned_Flat_Table", "Fact_Transactions", "Dim_Date", "Dim_MerchantCategory",
               "Dim_PaymentMethod", "Dim_Occupation", "Dim_Region", "Dim_Gender", "Data_Quality_Log"]

for name in data_sheets:
    ws = wb[name]
    ws.freeze_panes = "A2"
    max_col = ws.max_column
    for col in range(1, max_col+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 6001)):
        for cell in row:
            cell.font = body_font
    # autosize columns (approx, based on header + sample)
    for col in range(1, max_col+1):
        letter = get_column_letter(col)
        header_len = len(str(ws.cell(row=1, column=col).value or ""))
        sample_lens = [len(str(ws.cell(row=r, column=col).value or "")) for r in range(2, min(ws.max_row, 50)+1)]
        width = max([header_len] + sample_lens) + 3
        ws.column_dimensions[letter].width = min(max(width, 10), 40)

wb.save(path)
print("Formatting applied to:", data_sheets)
