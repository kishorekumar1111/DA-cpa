import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment

path = "/home/claude/day1/output/Master_Finance_Data_Cleaned.xlsx"
wb = openpyxl.load_workbook(path)

if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws = wb.create_sheet("Dashboard", 0)  # make it the first sheet

ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2

title_fill = PatternFill("solid", fgColor="2E5266")
ws["B2"] = "Day 1 — Data Integration & Preprocessing (ETL): Data Quality Dashboard"
ws["B2"].font = Font(size=16, bold=True, color="FFFFFF")
ws.merge_cells("B2:N2")
ws["B2"].fill = title_fill
ws["B2"].alignment = Alignment(vertical="center", indent=1)
ws.row_dimensions[2].height = 30

ws["B3"] = "Project: Algorithmic Credit Scoring Using Behavioral and Transactional Data | Source: Semi_Unclear_Finance_Dataset_5000.xlsx | Rows: 5,000"
ws["B3"].font = Font(size=10, italic=True, color="666666")
ws.merge_cells("B3:N3")

images = [
    ("01_missing_values.png", "B5"),
    ("02_region_before_after.png", "H5"),
    ("03_income_by_occupation.png", "B24"),
    ("04_default_rate_by_occupation.png", "H24"),
    ("05_monthly_volume.png", "B43"),
    ("06_credit_score_dist.png", "H43"),
]

for fname, cell in images:
    img = XLImage(f"/home/claude/day1/charts/{fname}")
    img.width = 430
    img.height = 260
    ws.add_image(img, cell)

for row_h in [5]:
    pass

wb.save(path)
print("Dashboard sheet added.")
print(wb.sheetnames)
