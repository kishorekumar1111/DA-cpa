import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment

path = "/home/claude/day2/output/Master_Finance_Data_EDA.xlsx"
wb = openpyxl.load_workbook(path)

if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws = wb.create_sheet("Dashboard", 0)
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2

title_fill = PatternFill("solid", fgColor="2E5266")
ws["B2"] = "Day 2 — Descriptive Statistics & EDA Dashboard"
ws["B2"].font = Font(size=16, bold=True, color="FFFFFF")
ws.merge_cells("B2:N2")
ws["B2"].fill = title_fill
ws["B2"].alignment = Alignment(vertical="center", indent=1)
ws.row_dimensions[2].height = 30

ws["B3"] = "Project: Algorithmic Credit Scoring | Source: Master_Finance_Data_Cleaned.xlsx (Day 1) | Rows analyzed: 5,000"
ws["B3"].font = Font(size=10, italic=True, color="666666")
ws.merge_cells("B3:N3")

images = [
    ("01_distributions_grid.png", "B5"),
    ("03_correlation_heatmap.png", "H5"),
    ("02_boxplots_outliers.png", "B33"),
    ("04_income_vs_creditscore.png", "H33"),
    ("05_default_rate_occ_region.png", "B52"),
    ("06_category_frequencies.png", "H52"),
]
for fname, cell in images:
    img = XLImage(f"/home/claude/day2/charts/{fname}")
    img.width = 430
    img.height = 260
    ws.add_image(img, cell)

wb.save(path)
print("Dashboard added.", wb.sheetnames)
